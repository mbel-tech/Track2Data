"""Tests for ExcelExporter."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from track2data.core.models import PreprocessReport
from track2data.exporters.base import ExportPayload
from track2data.exporters.excel import ExcelExporter

# Sheet names the exporter is documented (and, per the source, guaranteed) to write,
# in the exact order ExcelExporter.write() creates them.
EXPECTED_SHEET_NAMES = [
    "Fish by Frame",
    "Activity Summary",
    "Group Dynamics",
    "Zone Occupancy",
    "Quality",
]

# ── fixtures ──────────────────────────────────────────────────────────────────


def _make_fish_by_frame_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "session_id": ["sess1", "sess1", "sess1"],
            "individual_id": [0, 0, 1],
            "frame": [0, 1, 0],
            "x": [11.5, 22.5, 33.5],
            "y": [44.5, 55.5, 66.5],
        }
    )


def _make_individual_metrics() -> dict[str, pd.DataFrame]:
    return {
        "IL-1": pd.DataFrame(
            {
                "session_id": ["sess1", "sess1"],
                "individual_id": [0, 1],
                "mean_speed_cm_s": [1.25, 2.75],
            }
        )
    }


def _make_group_metrics() -> dict[str, pd.DataFrame]:
    return {
        "GL-1": pd.DataFrame(
            {
                "session_id": ["sess1"],
                "mean_NN_dist_cm": [9.5],
            }
        )
    }


def _make_zone_metrics() -> dict[str, pd.DataFrame]:
    return {
        "ZL-1": pd.DataFrame(
            {
                "session_id": ["sess1"],
                "zone_name": ["A"],
                "time_s": [12.0],
            }
        )
    }


def _make_diagnostic_metrics() -> dict[str, pd.DataFrame]:
    return {
        "D-1": pd.DataFrame(
            {
                "session_id": ["sess1"],
                "diagnostic": ["ok"],
                "value": [1.0],
            }
        )
    }


@pytest.fixture()
def minimal_payload() -> ExportPayload:
    return ExportPayload(
        session_id="sess1",
        project_name="TestProject",
        project_hash="abcd1234efgh5678",
        app_version="0.1.0",
        fish_by_frame=_make_fish_by_frame_df(),
        individual_metrics=_make_individual_metrics(),
        group_metrics=_make_group_metrics(),
        zone_metrics=_make_zone_metrics(),
        diagnostic_metrics=_make_diagnostic_metrics(),
        preprocess_report=PreprocessReport(),
        manifest_json=json.dumps({"schema_version": 1, "project_name": "TestProject"}),
    )


def _empty_metrics_payload() -> ExportPayload:
    return ExportPayload(
        session_id="sess1",
        project_name="EmptyProject",
        project_hash="a" * 16,
        app_version="0.1.0",
        fish_by_frame=_make_fish_by_frame_df(),
        individual_metrics={},
        group_metrics={},
        zone_metrics={},
        diagnostic_metrics={},
        preprocess_report=PreprocessReport(),
        manifest_json="{}",
    )


# ── ExcelExporter tests ──────────────────────────────────────────────────────


class TestExcelExporter:
    def test_exporter_name(self) -> None:
        assert ExcelExporter.name == "excel"

    def test_exporter_extension(self) -> None:
        assert ExcelExporter.file_extension == ".xlsx"

    def test_write_returns_single_path(
        self, tmp_path: Path, minimal_payload: ExportPayload
    ) -> None:
        exporter = ExcelExporter()
        paths = exporter.write(minimal_payload, tmp_path)
        assert isinstance(paths, list)
        assert len(paths) == 1

    def test_write_returns_xlsx_path(
        self, tmp_path: Path, minimal_payload: ExportPayload
    ) -> None:
        exporter = ExcelExporter()
        paths = exporter.write(minimal_payload, tmp_path)
        assert paths[0].suffix == ".xlsx"

    def test_xlsx_exists_on_disk(
        self, tmp_path: Path, minimal_payload: ExportPayload
    ) -> None:
        exporter = ExcelExporter()
        paths = exporter.write(minimal_payload, tmp_path)
        assert paths[0].exists()

    def test_xlsx_filename_includes_project_name(
        self, tmp_path: Path, minimal_payload: ExportPayload
    ) -> None:
        exporter = ExcelExporter()
        paths = exporter.write(minimal_payload, tmp_path)
        assert paths[0].name == "Track2Data_TestProject.xlsx"

    def test_workbook_sheet_names(
        self, tmp_path: Path, minimal_payload: ExportPayload
    ) -> None:
        exporter = ExcelExporter()
        paths = exporter.write(minimal_payload, tmp_path)
        wb = openpyxl.load_workbook(paths[0])
        assert wb.sheetnames == EXPECTED_SHEET_NAMES

    def test_fish_by_frame_sheet_header_row(
        self, tmp_path: Path, minimal_payload: ExportPayload
    ) -> None:
        exporter = ExcelExporter()
        paths = exporter.write(minimal_payload, tmp_path)
        wb = openpyxl.load_workbook(paths[0])
        ws = wb["Fish by Frame"]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert header == ["session_id", "individual_id", "frame", "x", "y"]

    def test_fish_by_frame_cell_values_match_payload(
        self, tmp_path: Path, minimal_payload: ExportPayload
    ) -> None:
        """A known value from fish_by_frame ends up in the corresponding cell."""
        exporter = ExcelExporter()
        paths = exporter.write(minimal_payload, tmp_path)
        wb = openpyxl.load_workbook(paths[0])
        ws = wb["Fish by Frame"]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        x_col = header.index("x") + 1
        y_col = header.index("y") + 1
        # index=False means row 2 (1-indexed) is the first data row.
        x_values = [ws.cell(row=r, column=x_col).value for r in (2, 3, 4)]
        y_values = [ws.cell(row=r, column=y_col).value for r in (2, 3, 4)]
        assert x_values == [11.5, 22.5, 33.5]
        assert y_values == [44.5, 55.5, 66.5]

    def test_activity_summary_sheet_contains_individual_metric_values(
        self, tmp_path: Path, minimal_payload: ExportPayload
    ) -> None:
        exporter = ExcelExporter()
        paths = exporter.write(minimal_payload, tmp_path)
        df = pd.read_excel(paths[0], sheet_name="Activity Summary")
        assert "mean_speed_cm_s" in df.columns
        assert sorted(df["mean_speed_cm_s"].tolist()) == [1.25, 2.75]

    def test_group_dynamics_sheet_contains_group_metric_values(
        self, tmp_path: Path, minimal_payload: ExportPayload
    ) -> None:
        exporter = ExcelExporter()
        paths = exporter.write(minimal_payload, tmp_path)
        df = pd.read_excel(paths[0], sheet_name="Group Dynamics")
        assert "mean_NN_dist_cm" in df.columns
        assert df["mean_NN_dist_cm"].tolist() == [9.5]

    def test_zone_occupancy_sheet_contains_zone_metric_values(
        self, tmp_path: Path, minimal_payload: ExportPayload
    ) -> None:
        exporter = ExcelExporter()
        paths = exporter.write(minimal_payload, tmp_path)
        df = pd.read_excel(paths[0], sheet_name="Zone Occupancy")
        assert "zone_name" in df.columns
        assert df["zone_name"].tolist() == ["A"]

    def test_quality_sheet_contains_diagnostic_metric_values(
        self, tmp_path: Path, minimal_payload: ExportPayload
    ) -> None:
        exporter = ExcelExporter()
        paths = exporter.write(minimal_payload, tmp_path)
        df = pd.read_excel(paths[0], sheet_name="Quality")
        assert "diagnostic" in df.columns
        assert df["diagnostic"].tolist() == ["ok"]

    def test_all_returned_paths_exist(
        self, tmp_path: Path, minimal_payload: ExportPayload
    ) -> None:
        exporter = ExcelExporter()
        paths = exporter.write(minimal_payload, tmp_path)
        for p in paths:
            assert p.exists(), f"{p} does not exist"

    def test_creates_out_dir_if_missing(
        self, tmp_path: Path, minimal_payload: ExportPayload
    ) -> None:
        out_dir = tmp_path / "nested" / "output"
        exporter = ExcelExporter()
        paths = exporter.write(minimal_payload, out_dir)
        assert paths[0].exists()

    def test_multi_entry_individual_metrics_are_merged(self, tmp_path: Path) -> None:
        """Two individual-metric frames sharing key columns outer-merge into one sheet."""
        payload = ExportPayload(
            session_id="sess1",
            project_name="MergeProject",
            project_hash="b" * 16,
            app_version="0.1.0",
            fish_by_frame=_make_fish_by_frame_df(),
            individual_metrics={
                "IL-1": pd.DataFrame(
                    {
                        "session_id": ["sess1", "sess1"],
                        "individual_id": [0, 1],
                        "mean_speed_cm_s": [1.25, 2.75],
                    }
                ),
                "IL-2": pd.DataFrame(
                    {
                        "session_id": ["sess1", "sess1"],
                        "individual_id": [0, 1],
                        "path_length_px": [100.0, 200.0],
                    }
                ),
            },
            group_metrics={},
            zone_metrics={},
            diagnostic_metrics={},
            preprocess_report=PreprocessReport(),
            manifest_json="{}",
        )
        exporter = ExcelExporter()
        paths = exporter.write(payload, tmp_path)
        df = pd.read_excel(paths[0], sheet_name="Activity Summary")
        assert "mean_speed_cm_s" in df.columns
        assert "path_length_px" in df.columns
        assert len(df) == 2

    def test_multi_entry_metrics_with_no_shared_columns_are_concatenated(
        self, tmp_path: Path
    ) -> None:
        """Two group-metric frames with disjoint columns fall into the concat branch
        (no shared key columns to merge on)."""
        payload = ExportPayload(
            session_id="sess1",
            project_name="ConcatProject",
            project_hash="c" * 16,
            app_version="0.1.0",
            fish_by_frame=_make_fish_by_frame_df(),
            individual_metrics={},
            group_metrics={
                "GL-1": pd.DataFrame({"mean_NN_dist_cm": [9.5]}),
                "GL-2": pd.DataFrame({"polarization": [0.42]}),
            },
            zone_metrics={},
            diagnostic_metrics={},
            preprocess_report=PreprocessReport(),
            manifest_json="{}",
        )
        exporter = ExcelExporter()
        paths = exporter.write(payload, tmp_path)
        df = pd.read_excel(paths[0], sheet_name="Group Dynamics")
        assert "mean_NN_dist_cm" in df.columns
        assert "polarization" in df.columns
        assert df["mean_NN_dist_cm"].tolist() == [9.5]
        assert df["polarization"].tolist() == [0.42]


class TestExcelExporterEmptyMetrics:
    """payload with empty individual_metrics={} etc. -- must still write a valid workbook."""

    def test_write_does_not_raise(self, tmp_path: Path) -> None:
        exporter = ExcelExporter()
        exporter.write(_empty_metrics_payload(), tmp_path)  # should not raise

    def test_write_returns_single_existing_xlsx(self, tmp_path: Path) -> None:
        exporter = ExcelExporter()
        paths = exporter.write(_empty_metrics_payload(), tmp_path)
        assert len(paths) == 1
        assert paths[0].exists()

    def test_workbook_is_openable_and_has_all_sheets(self, tmp_path: Path) -> None:
        exporter = ExcelExporter()
        paths = exporter.write(_empty_metrics_payload(), tmp_path)
        wb = openpyxl.load_workbook(paths[0])
        assert wb.sheetnames == EXPECTED_SHEET_NAMES

    def test_fish_by_frame_sheet_still_has_data(self, tmp_path: Path) -> None:
        """fish_by_frame is independent of the metric dicts and is always populated."""
        exporter = ExcelExporter()
        paths = exporter.write(_empty_metrics_payload(), tmp_path)
        df = pd.read_excel(paths[0], sheet_name="Fish by Frame")
        assert len(df) == 3

    def test_empty_metric_sheets_have_no_data_rows(self, tmp_path: Path) -> None:
        exporter = ExcelExporter()
        paths = exporter.write(_empty_metrics_payload(), tmp_path)
        for sheet in ("Activity Summary", "Group Dynamics", "Zone Occupancy", "Quality"):
            df = pd.read_excel(paths[0], sheet_name=sheet)
            assert len(df) == 0
